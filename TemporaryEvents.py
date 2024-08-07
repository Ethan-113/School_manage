from inspect import _void
import openpyxl
import logging
from pywebio.input import *
from pywebio.output import *

# 临时事务表的路径
file_path_root = '..\\临时事务表\\'

# 日志文件路径
log_file = '..\\Logs\\'

# 每个临时事务的描述
description_dict = {1: "酒店业务办理", 2: "一号食堂用餐", 3: "二号食堂用餐", 4: "ATM存取款", 5: "收发快递" \
    , 6: "参加学生活动", 7: "洗澡", 8: "喝咖啡", 9: "超市购物", 10: "图书馆读书或学习" \
    , 11: "拿外卖", 12: "健身", 13: "通话相关业务办理", 14: "理发", 15: "打印资料" \
    , 16: "体育馆打球", 17: "游泳", 18: "学生服务大厅办理业务", 19: "看病", 20: "操场活动"}

# 每个临时事务类型的名称
name_dict = {1: "hotel", 2: "dining_1", 3: "dining_2", 4: "ATM", 5: "express_stations", 6: "student_center",
             7: "bathhouse" \
    , 8: "coffee", 9: "supermaket", 10: "library", 11: "takeaway", 12: "gym", 13: "telecom", 14: "salon" \
    , 15: "print_shop", 16: "stadium", 17: "natatorium", 18: "service_lobby", 19: "hospital", 20: "playground"}

# 每个临时事务类型的地点坐标
place_dict = {1: [7, 3], 2: [11, 22], 3: [7, 7], 4: [19, 1], 5: [2, 10], 6: [7, 12], 7: [7, 16], 8: [9, 12], 9: [7, 17],
              10: [14, 15] \
    , 11: [3, 25], 12: [3, 22], 13: [3, 27], 14: [11, 24], 15: [11, 27], 16: [14, 20], 17: [17, 29], 18: [3, 20],
              19: [27, 22], 20: [26, 25]}

# 每周开始的日期
week_date = {1: 305, 2: 312, 3: 319, 4: 326, 5: 402, 6: 409, 7: 416, 8: 423,
             9: 430, 10: 507, 11: 514, 12: 521, 13: 528, 14: 604, 15: 611,
             16: 618, 17: 625}

# 用于存储事件
events_list = []

# 返回有闹钟的事件
clock_list = []

# 每节课开始的时间
time_start = {1: 8.00, 2: 9.00, 3: 10.00, 4: 13.00, 5: 14.00, 6: 15.00, 7: 18.00, 8: 19.00}

# 每节课结束的时间
time_end = {1: 9.00, 2: 10.00, 3: 11.00, 4: 14.00, 5: 15.00, 6: 16.00, 7: 19.00, 8: 20.00}

# 记录每张课表的路径
stu_dir = {1: '..\\课程相关表\\stu1.xlsx',
           2: '..\\课程相关表\\stu2.xlsx'}

# 记录sheet名称
week_dir = {1: '1', 2: '2', 3: '3', 4: '4' \
    , 5: '5', 6: '6', 7: '7', 8: '8' \
    , 9: '9', 10: '10', 11: '11', 12: '12' \
    , 13: '13', 14: '14', 15: '15', 16: '16'}
global num_class, week, day


# 事件类型，名称与坐标用字典存储
class Event:
    def __init__(self, type, time_hour, time_min, clock):
        self.type = type
        self.time_hour = time_hour
        self.time_min = time_min
        self.place = place_dict[type]
        self.name = name_dict[type]
        self.description = description_dict[type]
        self.clock = clock

    def getClock(self):
        return self.clock

    def getDescription(self):
        return self.description

    def getType(self):
        return self.type

    def getPlace(self):
        return self.place

    def getName(self):
        return self.name

    def getHour(self):
        return self.time_hour

    def getMin(self):
        return self.time_min

    def forCompare(self):
        compare = self.time_hour + self.time_min / 100
        round(compare, 2)
        return compare


# 将临时事件添加到Excel表格中
def addEvent(Event):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook['Sheet1']
    max = worksheet.max_row

    # 每次添加时将要添加的事件时间与表中已有事件的时间进行比较，从而实现排序，时间早的在上晚的在下（类似插入排序）
    for i in range(max, 1, -1):
        if Event.forCompare() < worksheet.cell(row=i, column=7).value:
            worksheet.cell(row=i + 1, column=1).value = worksheet.cell(row=i, column=1).value
            worksheet.cell(row=i + 1, column=2).value = worksheet.cell(row=i, column=2).value
            worksheet.cell(row=i + 1, column=3).value = worksheet.cell(row=i, column=3).value
            worksheet.cell(row=i + 1, column=4).value = worksheet.cell(row=i, column=4).value
            worksheet.cell(row=i + 1, column=5).value = worksheet.cell(row=i, column=5).value
            worksheet.cell(row=i + 1, column=6).value = worksheet.cell(row=i, column=6).value
            worksheet.cell(row=i + 1, column=7).value = worksheet.cell(row=i, column=7).value
            worksheet.cell(row=i + 1, column=8).value = worksheet.cell(row=i, column=8).value
        else:
            break
    worksheet.cell(row=i + 1, column=1).value = Event.getType()
    worksheet.cell(row=i + 1, column=2).value = Event.getName()
    worksheet.cell(row=i + 1, column=3).value = Event.getHour()
    worksheet.cell(row=i + 1, column=4).value = Event.getMin()
    worksheet.cell(row=i + 1, column=5).value = Event.getPlace()[0]
    worksheet.cell(row=i + 1, column=6).value = Event.getPlace()[1]
    worksheet.cell(row=i + 1, column=7).value = Event.forCompare()
    worksheet.cell(row=i + 1, column=8).value = Event.getClock()
    workbook.save(file_path)


# 将临时事务从表中删除，num表示行标
def deleteEvent(num, stu_num):
    global file_path_now
    file_path_now = file_path_root + 'Temporary' + str(stu_num) + '.xlsx'
    workbook = openpyxl.load_workbook(file_path_now)
    worksheet = workbook['Sheet1']
    worksheet.delete_rows(num)
    workbook.save(file_path_now)


def temporary_deleteAll(stu_num):
    global file_path_root, file_path_now
    file_path_now = file_path_root + 'Temporary' + str(stu_num) + '.xlsx'
    workbook = openpyxl.load_workbook(file_path_now)
    worksheet = workbook['Sheet1']
    max = worksheet.max_row
    for i in range(max, 2, -1):
        deleteEvent(i, stu_num)


# 将表格中的所有事件打印出来
def showAllEvent():
    i = 1
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook['Sheet1']
    max = worksheet.max_row
    with popup(title='全部事件'):
        for row in worksheet.iter_rows(min_row=3, max_row=max, values_only=True):
            put_text("事件编号为[" + str(i) + "]，事件类型为：" + str(row[0]) + "  事件名称为：" + row[
                1] + "  事件时间为：" + str(row[2]) + ":" + str(row[3]) + "  事件地点为：" + str(row[4]) + "," + str(
                row[5]))
            i = i + 1


# 读取表格中的事件并存入列表中，从而实现存储
def loadEvent():
    # 读取前保证列表为空
    events_list.clear()
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook['Sheet1']
    max = worksheet.max_row

    # 表中特定元素进行事件创建
    for row in worksheet.iter_rows(min_row=3, max_row=max, values_only=True):
        type = row[0]
        time_hour = row[2]
        time_min = row[3]
        clock = row[7]
        event = Event(type, time_hour, time_min, clock)
        events_list.append(event)


def temporary_forevent(stu_num):
    global file_path_root
    file_path_now = file_path_root + 'Temporary' + str(stu_num) + '.xlsx'
    clock_list.clear()
    workbook = openpyxl.load_workbook(file_path_now)
    worksheet = workbook['Sheet1']
    max = worksheet.max_row

    for row in worksheet.iter_rows(min_row=3, max_row=max, values_only=True):
        if row[7] == 1:
            type = row[0]
            time_hour = row[2]
            time_min = row[3]
            clock = row[7]
            event = Event(type, time_hour, time_min, clock)
            clock_list.append(event)
    return clock_list


# 二分法得到目标最左的数组下标
def low_bound(list, x, y, v):
    while x < y:
        m = int((x + y) / 2)
        if list[m].forCompare() >= v:
            y = m
        else:
            x = m + 1
    return x


# 二分法得到目标最右的数组下标
def up_bound(list, x, y, v):
    while x < y:
        m = int((x + y) / 2)
        if list[m].forCompare() <= v:
            x = m + 1
        else:
            y = m
    return x


# 查询目标时间的事件所在的范围，如果左右下标相等说明目标不存在，精确到分钟的精确查找
def searchEventAccurate(targetTime_hour, targetTime_min):
    targetTime = targetTime_hour + targetTime_min / 100
    # 调用函数得到最左下标
    L = low_bound(events_list, 0, len(events_list), targetTime)
    # 调用函数得到最右下标
    R = up_bound(events_list, 0, len(events_list), targetTime)
    if L == R:
        with popup(title='目标临时事件'):
            put_text("目标时间不存在临时事件安排")
    else:
        with popup(title='目标临时事件'):
            for i in range(L, R):
                put_text(str(events_list[i].getHour()) + '时' + str(events_list[i].getMin()) + '分到' + str(
                    events_list[i].getPlace()) + '进行' + str(events_list[i].getName()) + '事件')


# 二分法得到目标最左的数组下标
def low_bound_about(list, x, y, v):
    while x < y:
        m = int((x + y) / 2)
        if list[m].getHour() >= v:
            y = m
        else:
            x = m + 1
    return x


# 二分法得到目标最右的数组下标
def up_bound_about(list, x, y, v):
    while x < y:
        m = int((x + y) / 2)
        if list[m].getHour() <= v:
            x = m + 1
        else:
            y = m
    return x


# 查询目标时间的事件所在的范围，如果左右下标相等说明目标不存在，精确到小时的粗查找
def searchEventAbout(targetTime):
    # 调用函数得到最左下标
    L = low_bound_about(events_list, 0, len(events_list), targetTime)
    # 调用函数得到最右下标
    R = up_bound_about(events_list, 0, len(events_list), targetTime)
    if L == R:
        with popup(title='目标临时事件'):
            put_text("目标时间不存在临时事件安排")
    else:
        with popup(title='目标临时事件'):
            for i in range(L, R):
                put_text(str(events_list[i].getHour()) + '时' + str(events_list[i].getMin()) + '分到' + str(
                    events_list[i].getPlace()) + '进行' + str(events_list[i].getName()) + '事件')


# 用事件类型进行搜索
def searchType(type):
    for i in range(len(events_list)):
        with popup(title='目标临时事件'):
            if type == events_list[i].getType():
                put_text(str(events_list[i].getHour()) + '时' + str(events_list[i].getMin()) + '分到' + str(
                    events_list[i].getPlace()) + '进行' + str(events_list[i].getName()) + '事件')


# 判断是否与课程冲突
def crashORnot(e):
    global num_class, week, day
    list_today = []
    workbook = openpyxl.load_workbook(stu_dir[num_class])
    worksheet = workbook[week_dir[week]]
    max = worksheet.max_row
    for row in worksheet.iter_rows(min_row=2, max_row=max, values_only=True):
        if row[day] != '\\':
            list_today.append(row[0])

    flag = 0
    time = e.forCompare()
    for i in list_today:
        start = time_start[i]
        end = time_end[i]
        if ((time >= start) & (time <= end)):
            flag = 1
            break
    if flag == 1:  # 与当天课程有冲突
        return False
    else:  # 无冲突
        return True


# 获取今天是第几周
def NOofweek(month_today, day_today):
    compare = month_today * 100 + day_today
    for i in range(1, 17):
        if compare >= week_date[i] and compare < week_date[i + 1]:
            return i


def button_clicked(btn):
    global stu_num2

    if btn == '添加临时事件':  # 添加
        fmt = "{:25}\t{:25}\t{:25}"
        with popup(title='提供的所有临时事件'):
            put_text(fmt.format("类型编号", "类型", "事件描述"))
            for i in range(1, 21):
                put_text(fmt.format(str(i), name_dict[i], description_dict[i]))
        name = int(input("输入事件类型编号："))
        hour = int(input("事件在多少时："))
        min = int(input("事件在多少分："))
        clock = int(input("是否设定闹钟，输入1为设定："))
        event = Event(name, hour, min, clock)
        if crashORnot(event):
            addEvent(event)
            loadEvent()
            logging.info('用户添加事件，%d时%d分，事件编号为%d', hour, min, name)
        else:
            with popup(title='有时间冲突'):
                put_text("与课程或活动时间冲突")

    elif btn == '删除事件':  # 删除
        showAllEvent()
        todelete = int(input("输入您想要删除的任务的编号："))
        deleteEvent(todelete + 2, stu_num2)
        loadEvent()
        toast('删除成功')
        logging.info('用户删除临时事件')

    elif btn == '查询事件':
        flag_search = int(input("按分钟的精确查询请输入1，或按小时的粗查询请输入2，按事件类型查询请输入3："))
        if flag_search == 1:
            hour_aim = int(input("所查询事件的小时："))
            min_aim = int(input("所查询事件的分钟："))
            searchEventAccurate(hour_aim, min_aim)
            # 记录日志
            logging.info('用户查询%d时%d分的事件', hour_aim, min_aim)

        elif flag_search == 2:
            hour_aim = int(input("所查询事件的小时："))
            searchEventAbout(hour_aim)
            # 记录日志
            logging.info('用户查询%d时的事件', hour_aim)

        elif flag_search == 3:
            type = int(input("输入事件类型："))
            searchType(type)
            # 记录日志
            logging.info('用户查询%d类型的事件', type)
        else:
            toast('类型不存在')

    elif btn == '显示所有事件':
        showAllEvent()
        # 记录日志
        logging.info('用户查看所有事件')


# 主操作模块
def temp_event_init(stu_num, stu_class, month_today, day_today, week_today):
    global num_class, week, day, log_file, file_path, file_path_root, stu_num2
    stu_num2 = stu_num
    num_class = stu_class
    # 获取今天是第几周
    week = NOofweek(month_today, day_today)
    # 获取今天是周几
    day = week_today

    file_path = file_path_root + 'Temporary' + str(stu_num) + '.xlsx'
    log_file = log_file + "Log-" + str(stu_num) + ".txt"
    logging.basicConfig(filename=log_file, level=logging.DEBUG, format='%(asctime)s -- %(message)s')

    # 首先加载事件表
    loadEvent()
    toast('临时事务模块')

    put_buttons(['添加临时事件', '删除事件', '查询事件', '显示所有事件'], onclick=button_clicked)