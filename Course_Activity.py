import os
import openpyxl
import pandas as pd
import decimal
import time
import logging
from pywebio.input import input
from pywebio.output import put_text, put_buttons, use_scope, put_markdown, toast, popup

path = '..\\课程相关表'
# 日志文件路径
log_file = '..\\Logs\\'
os.chdir(path)
file_path1 = r'BIU.xlsx'
file_path2 = r'stu1.xlsx'
file_path3 = r'stu2.xlsx'
df1 = pd.read_excel(file_path1, sheet_name=None)
df2 = pd.read_excel(file_path2, sheet_name=None)
df3 = pd.read_excel(file_path3, sheet_name=None)
week_date = {1: 20230305, 2: 20230312, 3: 20230319, 4: 20230326, 5: 20230402, 6: 20230409, 7: 20230416, 8: 20230423,
             9: 20230430, 10: 20230507, 11: 20230514, 12: 20230521, 13: 20230528, 14: 20230604, 15: 20230611,
             16: 20230618, 17: 20230625}
time_start = {1: 800, 2: 900, 3: 1000, 4: 1300, 5: 1400, 6: 1500, 7: 1800, 8: 1900}
time_start_str = {1: '08:00', 2: '09:00', 3: '10:00', 4: '13:00', 5: '14:00', 6: '15:00', 7: '18:00', 8: '19:00'}
student_data = df1['student'].to_dict(orient='records')
""" BIU student sheet数据变成字典 """
course_data = df1['course'].to_dict(orient='records')
""" BIU course sheet数据变成字典 """
instructor_data = df1['instructor'].to_dict(orient='records')
infrastructure_data = df1['infrastructure'].to_dict(orient='records')
set_count = 0
""" #用来记录更改课程表的操作次数 """
workbook1 = openpyxl.load_workbook(file_path1)
workbook2 = openpyxl.load_workbook(file_path2)
workbook3 = openpyxl.load_workbook(file_path3)
student_sheet = workbook1['student']
course_sheet = workbook1['course']
infra_sheet = workbook1['infrastructure']
instructor_sheet = workbook1['instructor']
second_day_data = {}
identity_data = {}


def querycourse():  # 查询课表
    querycourse_inf = input("请输入你需要查询的周或课程名称或地址(输入Exit退出该操作)：")
    count = 0
    if querycourse_inf == 'Exit':
        return None
    elif querycourse_inf.isdigit():
        if decimal.Decimal(querycourse_inf) > 16:
            toast("本学期最多16周")
            count += 1
            querycourse()
        else:
            logging.info('用户查询第%d周课表', int(querycourse_inf))
            if identity_data['class'] == 202101:
                week = int(querycourse_inf)
                week_course = pd.read_excel(file_path2, sheet_name=week - 1)
                with popup(title='课表'):
                    put_text(week_course)
            elif identity_data['class'] == 202102:
                week = int(querycourse_inf)
                week_course = pd.read_excel(file_path3, sheet_name=week - 1)
                with popup(title='课表'):
                    put_text(week_course)
            count += 1
            querycourse()
    else:
        for course in course_data:  # 以上都不是，就判断输入的是课程名称
            if course['course_name'] == querycourse_inf:
                logging.info('用户查询%s课程', str(querycourse_inf))
                with popup(title='课程信息'):
                    put_text(course)
                count += 1
                querycourse()
                break
        if count == 0:
            for course in course_data:
                if course['course_spot'] == querycourse_inf:
                    with popup(title='课程信息'):
                        put_text(course)
                    count += 1
            if count != 0:
                logging.info('用户查询在%s上课的课程', str(querycourse_inf))
                querycourse()
    if count == 0:  # 如果count为0，说明没有查询到课程
        toast('输入错误，没有查询到相关信息，请重新输入')
        querycourse()


def identity():  # 判断账号密码是否符合，登录
    account = input("请输入你的账号：")
    secret = input("请输入你的密码：")
    for stu in student_data:
        if str(account) == str(stu['stu_num']) and str(secret) == str(stu['stu_secret']):
            return stu  # 是学生则返回存储学生所有信息的字典
    for instructor in instructor_data:
        if str(account) == str(instructor['ins_account']) and str(secret) == str(instructor['ins_secret']):
            return instructor  # 是辅导员返回存储辅导员所有信息的字典
    end = input("账号或密码输入错误，请选择是否重新输入（输入Exit结束，其他输入将继续进行）：")
    if end == 'Exit':
        return 0
    else:
        return identity()


def IS_crash(name, week, info1, info2, IS_exam):  # 查询课程设置时间是否与其他课程冲突
    if identity_data['class'] == 202101:
        week_timetable = workbook2[workbook2.sheetnames[int(week) - 1]]
        if week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value != "\\":
            put_text('该时间已有课程')
            put_text(week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value)
            return 1
        if IS_exam == 0:
            week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value = name
        else:
            week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value = str(
                name) + ".exam"
        workbook2.save(file_path2)
    elif identity_data['class'] == 202102:
        week_timetable = workbook2[workbook3.sheetnames[int(week) - 1]]
        if week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value != "\\":
            put_text('该时间已有课程')
            put_text(week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value)
            return 1
        if IS_exam == 0:
            week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value = name
        else:
            week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value = str(
                name) + ".exam"
        workbook3.save(file_path3)
    return 0


def time_set(week, name, rrow, cclumn):  # 课程设置时间保存
    info1 = input("输入课程设置时间的星期几")
    while info1.isdigit() != True or decimal.Decimal(info1) <= 0 or decimal.Decimal(info1) > 7:
        info1 = input("输入不符合，请重新输入")
    info2 = input("输入课程设置时间的第几节课，一天最多八节")
    while info2.isdigit() != True or decimal.Decimal(info2) <= 0 or decimal.Decimal(info2) > 8:
        info2 = input("输入不符合，请重新输入")
    if IS_crash(name, week, info1, info2, 0) == 0:
        course_sheet.cell(row=rrow, column=cclumn).value = str(
            course_sheet.cell(row=rrow, column=cclumn).value) + "(" + str(info1) + "," + str(info2) + ")"
        workbook1.save(file_path1)
    select = input("是否继续设置该周的时间，输入Go继续，其他结束")
    if select == "Go":
        time_set(week, name, rrow, cclumn)
    else:
        return None


def delete_timetable(week, info1, info2):  # 删除课程表上原来的课程时间
    toast(info1)
    if identity_data['class'] == 202101:
        week_timetable = workbook2[workbook2.sheetnames[int(week) - 1]]
        week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value = "\\"
        workbook2.save(file_path2)
    elif identity['class'] == 202102:
        week_timetable = workbook2[workbook3.sheetnames[int(week) - 1]]
        week_timetable.cell(row=decimal.Decimal(info2) + 1, column=decimal.Decimal(info1) + 1).value = "\\"
        workbook3.save(file_path3)


def delete_time(week, time):  # 删除课程信息里的课程时间
    L_num = 0
    R_num = 0
    Dou_num = 0
    if time != "":
        for i in range(len(str(time))):
            if str(time)[i] == "(": L_num = i
            if str(time)[i] == ",": Dou_num = i
            if str(time)[i] == ")": R_num = i
            if L_num < Dou_num and Dou_num < R_num:
                info1 = str(time)[L_num + 1:Dou_num]
                info2 = str(time)[Dou_num + 1:R_num]
                delete_timetable(week, info1, info2)


def course_row(course_name):  # 获取到需要进行更改的行
    Min_row = course_sheet.min_row
    Max_row = course_sheet.max_row
    mid = int((Min_row + Max_row) / 2)
    row_course = 0
    while 1:
        if course_sheet.cell(row=mid, column=6).value == identity_data['class']:
            mid_up = mid
            mid_down = mid
            while 1:
                if course_sheet.cell(row=mid_up, column=1).value == course_name and course_sheet.cell(row=mid,
                                                                                                      column=6).value == \
                        identity_data['class']:
                    row_course = mid_up
                    break
                elif course_sheet.cell(row=mid_up, column=6).value != identity_data['class']:
                    break
                mid_up = mid_up - 1
            if row_course == 0:
                while 1:
                    if course_sheet.cell(row=mid_down, column=1).value == course_name and course_sheet.cell(row=mid,
                                                                                                            column=6).value == \
                            identity_data['class']:
                        row_course = mid_down
                        break
                    elif course_sheet.cell(row=mid_down, column=6).value != identity_data['class']:
                        row_course = mid_down
                        break
                    mid_down += 1
            break
        if int(course_sheet.cell(row=mid, column=6).value) < int(identity_data['class']):
            Min_row = mid
            mid = int((Min_row + Max_row) / 2)
        if int(course_sheet.cell(row=mid, column=6).value) > int(identity_data['class']):
            Max_row = mid
            mid = int((Min_row + Max_row) / 2)
    return row_course


def delete_examtime(examtime):  # 删除考试的时间
    L_num = 0
    R_num = 0
    Dou_L_num = 0
    Dou_R_num = 0
    if examtime != "":
        for i in range(len(str(examtime))):
            if str(examtime)[i] == "(": L_num = i
            if str(examtime)[i] == ",":
                if Dou_L_num <= Dou_R_num:
                    Dou_L_num = i
                elif Dou_R_num < Dou_L_num:
                    Dou_R_num = i
            if str(examtime)[i] == ")": R_num = i
            if L_num < Dou_L_num and Dou_L_num < Dou_R_num and Dou_R_num < R_num:
                week = str(examtime)[L_num + 1:Dou_L_num]
                info1 = str(examtime)[Dou_L_num + 1:Dou_R_num]
                info2 = str(examtime)[Dou_R_num + 1:R_num]
                delete_timetable(week, info1, info2)


def examtime_set(name, rrow, cclumn):  # 设置考试的时间
    week = input("输入考试时间的第几周,该学期共16周")
    while week.isdigit() != True or decimal.Decimal(week) <= 0 or decimal.Decimal(week) > 16:
        week = input("输入不符合，请重新输入")
    info1 = input("输入课程设置时间的星期几")
    while info1.isdigit() != True or decimal.Decimal(info1) <= 0 or decimal.Decimal(info1) > 7:
        info1 = input("输入不符合，请重新输入")
    info2 = input("输入课程设置时间的第几节课，一天最多八节")
    while info2.isdigit() != True or decimal.Decimal(info2) <= 0 or decimal.Decimal(info2) > 8:
        info2 = input("输入不符合，请重新输入")
    if IS_crash(name, week, info1, info2, 1) == 0:
        course_sheet.cell(row=rrow, column=cclumn).value = str(
            course_sheet.cell(row=rrow, column=cclumn).value) + "(" + str(week) + "," + str(info1) + "," + str(
            info2) + ")"
        workbook1.save(file_path1)
    select = input("是否继续设置考试时间，Go继续")
    if select == "Go":
        examtime_set(name, rrow, cclumn)
    else:
        return None


def set_timetable():  # 通过调用以上函数实现对课程表的更改
    i = 0
    for key in identity_data.keys():
        if key == 'ins_account':
            i += 1
            break
    if i == 0:
        toast("您没有权限更改课程。")
        return None
    set_course_name = input("请输入你要设置的课程名称")
    row_course = course_row(set_course_name)
    while True:
        if course_sheet.cell(row=row_course, column=6).value == identity_data['class']:
            IS_new = 0
            break
        if course_sheet.cell(row=row_course, column=6).value != identity_data['class']:
            IS_new = input("是否为新增课程,输入1是，输入Exit退出设置，其他不是")
            if IS_new == "1":
                break
            elif IS_new == "Exit":
                return None
        set_course_name = input("请输入你要设置的课程名称")
        row_course = course_row(set_course_name)
    if IS_new == 0:
        toast("查询到有该课程")
        data = df1["course"]
        put_text(data.loc[row_course - 2])
        logging.info('用户更改%s课程', str(set_course_name))
        i = 2
        while i <= course_sheet.max_column:
            if i != 6:
                put_text(course_sheet.cell(row=1, column=i).value)
            if i == 2:
                IS_in = input("是否进入修改，回车键进入下一项,其他任意键进入修改")
            elif i > 2 and i != 6:
                IS_in = input("是否进入修改，回车键进入下一项,-1进入上一级修改,其他任意键进入修改")
            if IS_in == "":
                i += 1
                continue
            elif IS_in == "-1" and i > 2 and i != 7:
                i -= 1
                continue
            elif IS_in == "-1" and i == 7:
                i -= 2
                continue
            if i == 2:
                for fra in infrastructure_data:
                    put_text(fra['infra_name'])
                info = input("请输入地址")
                for infra in range(1, infra_sheet.max_row + 2):
                    if infra <= infra_sheet.max_row:
                        if infra_sheet.cell(row=infra, column=2).value == info:
                            course_sheet.cell(row=row_course, column=i).value = info
                            break
                while infra > infra_sheet.max_row:
                    info = input("请重新输入：")
                    for infra in range(1, infra_sheet.max_row + 2):
                        if infra <= infra_sheet.max_row:
                            if infra_sheet.cell(row=infra, column=2).value == info:
                                course_sheet.cell(row=row_course, column=i).value = info
                                break
                logging.info('用户将%s课程的上课地点改为%s', str(set_course_name), str(info))
                workbook1.save(file_path1)
            elif i == 4:
                if str(course_sheet.cell(row=row_course, column=3).value) != "":
                    if course_sheet.cell(row=row_course, column=3).value:
                        for fra in infrastructure_data:
                            put_text(fra['infra_name'])
                        info = input("请输入考试地点")
                        for infra in range(1, infra_sheet.max_row + 2):
                            if infra <= infra_sheet.max_row:
                                if infra_sheet.cell(row=infra, column=2).value == info:
                                    course_sheet.cell(row=row_course, column=i).value = info
                                    break
                        while infra > infra_sheet.max_row:
                            info = input("请重新输入：")
                            for infra in range(1, infra_sheet.max_row + 2):
                                if infra <= infra_sheet.max_row:
                                    if infra_sheet.cell(row=infra, column=2).value == info:
                                        course_sheet.cell(row=row_course, column=i).value = info
                                        break
                logging.info('用户将%s课程的考试地点改为%s', str(set_course_name), str(info))
                workbook1.save(file_path1)
            elif i == 3:
                delete_examtime(course_sheet.cell(row=row_course, column=i).value)
                course_sheet.cell(row=row_course, column=i).value = ""
                workbook1.save(file_path1)
                examtime_set(set_course_name, row_course, i)
                logging.info('用户将%s课程考试时间改为%s', str(set_course_name),
                             str(course_sheet.cell(row=row_course, column=i).value))
            elif i == 5:
                info = input("请输入老师姓名")
                course_sheet.cell(row=row_course, column=i).value = info
                logging.info('用户将%s课程的任课老师改为%s', str(set_course_name), str(info))
                workbook1.save(file_path1)
            elif i == 7:
                info = input("是否为线上Yes or No")
                while info != "Yes" and info != "No":
                    info = input("Yes or No")
                course_sheet.cell(row=row_course, column=i).value = info
                if info == "Yes":
                    logging.info('用户将%s课程改为线上上课', str(set_course_name))
                    course_sheet.cell(row=row_course, column=i - 1).value = '\\'
                    logging.info('用户更改的%s课程为线上进行,上课地点改为空', str(set_course_name))
                else:
                    logging.info('用户将%s课程改为线下上课', str(set_course_name))
                workbook1.save(file_path1)
            elif i == 8:
                if course_sheet.cell(row=row_course, column=7).value == "No":
                    course_sheet.cell(row=row_course, column=i).value = "\\"
                else:
                    info = input("请输入上课链接")
                    course_sheet.cell(row=row_course, column=i).value = info
                    logging.info('用户将%s课程的上课链接改为%s', str(set_course_name), str(info))
                workbook1.save(file_path1)
            elif i >= 9 and i <= 24:
                week = course_sheet.cell(row=1, column=i).value
                delete_time(week, course_sheet.cell(row=row_course, column=i).value)
                course_sheet.cell(row=row_course, column=i).value = ""
                workbook1.save(file_path1)
                time_set(week, set_course_name, row_course, i)
                logging.info('用户将%s课程第%d周上课时间改为%s', str(set_course_name), i - 8,
                             str(course_sheet.cell(row=row_course, column=i).value))
            i += 1
    elif IS_new == "1":
        course_sheet.insert_rows(row_course)
        workbook1.save(file_path1)
        logging.info('用户增加%s课程', str(set_course_name))
        i = 1
        while i <= course_sheet.max_column:
            if i > 2:
                if IS_in == "-1" and i > 2 and i != 8 and i != 24:
                    i -= 2
                elif IS_in == "-1" and i == 8 and i != 24:
                    i -= 3
                elif IS_in == "1":
                    i -= 1
            if i > 1 and i != 6:
                put_text(course_sheet.cell(row=1, column=i).value)
            if i == 1:
                course_sheet.cell(row=row_course, column=i).value = set_course_name
                workbook1.save(file_path1)
            elif i == 2:
                for fra in infrastructure_data:
                    put_text(fra['infra_name'])
                info = input("请输入课程地点")
                for infra in range(1, infra_sheet.max_row + 2):
                    if infra <= infra_sheet.max_row:
                        if infra_sheet.cell(row=infra, column=2).value == info:
                            course_sheet.cell(row=row_course, column=i).value = info
                            break
                while infra > infra_sheet.max_row:
                    info = input("请重新输入：")
                    for infra in range(1, infra_sheet.max_row + 2):
                        if infra <= infra_sheet.max_row:
                            if infra_sheet.cell(row=infra, column=2).value == info:
                                course_sheet.cell(row=row_course, column=i).value = info
                                break
                logging.info('用户将%s课程的上课地点设为%s', str(set_course_name), str(info))
                workbook1.save(file_path1)
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            elif i == 4:
                if course_sheet.cell(row=row_course, column=3).value:
                    for fra in infrastructure_data:
                        put_text(fra['infra_name'])
                    info = input("请输入考试地点")
                    for infra in range(1, infra_sheet.max_row + 2):
                        if infra <= infra_sheet.max_row:
                            if infra_sheet.cell(row=infra, column=2).value == info:
                                course_sheet.cell(row=row_course, column=i).value = info
                                break
                    while infra > infra_sheet.max_row:
                        info = input("请重新输入：")
                        for infra in range(1, infra_sheet.max_row + 2):
                            if infra <= infra_sheet.max_row:
                                if infra_sheet.cell(row=infra, column=2).value == info:
                                    course_sheet.cell(row=row_course, column=i).value = info
                                    break
                logging.info('用户将%s课程的考试地点设为%s', str(set_course_name), str(info))
                workbook1.save(file_path1)
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            elif i == 3:
                IS_set = input("输入1表示需要设置考试时间，其他表示不需要")
                if IS_set == '1':
                    course_sheet.cell(row=row_course, column=i).value = ""
                    workbook1.save(file_path1)
                    week = course_sheet.cell(row=1, column=i).value
                    examtime_set(set_course_name, row_course, i)
                logging.info('用户将%s课程考试时间设为%s', str(set_course_name),
                             str(course_sheet.cell(row=row_course, column=i).value))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            elif i == 5:
                info = input("请输入老师信息")
                while info == "":
                    info = input("输入老师信息")
                course_sheet.cell(row=row_course, column=i).value = info
                workbook1.save(file_path1)
                logging.info('用户将%s课程的任课老师设为%s', str(set_course_name), str(info))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            elif i == 6:
                course_sheet.cell(row=row_course, column=i).value = identity_data['class']
                workbook1.save(file_path1)
            elif i == 7:
                info = input("请输入是否为线上Yes or No")
                while info != "Yes" and info != "No":
                    info = input("Yes or No")
                course_sheet.cell(row=row_course, column=i).value = info
                workbook1.save(file_path1)
                if info == "Yes":
                    logging.info('用户将%s课程设为线上上课', str(set_course_name))
                    course_sheet.cell(row=row_course, column=i - 1).value = '\\'
                    logging.info('用户设置的%s课程为线上进行,上课地点改为空', str(set_course_name))
                else:
                    logging.info('用户将%s课程设为线下上课', str(set_course_name))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            elif i == 8:
                if course_sheet.cell(row=row_course, column=7).value == "No":
                    put_text("线上课程，已自动将上课链接设置为空")
                    course_sheet.cell(row=row_course, column=i).value = "\\"
                else:
                    info = input("请输入上课链接")
                    course_sheet.cell(row=row_course, column=i).value = info
                workbook1.save(file_path1)
                logging.info('用户将%s课程的上课链接设为%s', str(set_course_name), str(info))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            elif i >= 9 and i <= 24:
                IS_set = input("输入1表示需要设置上课时间，其他表示不需要")
                if IS_set == '1':
                    course_sheet.cell(row=row_course, column=i).value = ""
                    workbook1.save(file_path1)
                    week = course_sheet.cell(row=1, column=i).value
                    time_set(week, set_course_name, row_course, i)
                logging.info('用户将%s课程第%d周上课时间设为%s', str(set_course_name), i - 8,
                             str(course_sheet.cell(row=row_course, column=i).value))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
                if i == 24:
                    if IS_in == '1':
                        i -= 1
                    elif IS_in == '-1':
                        i -= 2
                    IS_in = '0'
            i += 1
    workbook1.save(file_path1)
    workbook2.save(file_path2)
    workbook3.save(file_path3)


def student_row():  # 获取到用户在用户信息表的第几行
    Min_row = student_sheet.min_row
    Max_row = student_sheet.max_row
    mid = int((Min_row + Max_row) / 2)
    row_student = 0
    while 1:
        if int(student_sheet.cell(row=mid, column=1).value) < identity_data['stu_num']:
            Min_row = mid
            mid = int((Min_row + Max_row) / 2)
        elif int(student_sheet.cell(row=mid, column=1).value) > identity_data['stu_num']:
            Max_row = mid
            mid = int((Min_row + Max_row) / 2)
        elif int(student_sheet.cell(row=mid, column=1).value) == identity_data['stu_num']:
            row_student = mid
            break
    return row_student


def get_crashclassnum(time):  # 获取到可能会与活动有冲突的第几节课
    class_num = []
    if int(time) + 100 > time_start[1] and int(time) + 100 < time_start[1] + 100:
        class_num.append(1)
    elif int(time) > time_start[8] and int(time) < time_start[8] + 100:
        class_num.append(8)
    for i in time_start:
        if i < 8:
            if int(time) >= time_start[i] and int(time) <= time_start[i + 1]:
                if int(time) < time_start[i] + 100 and int(time) + 100 <= time_start[i + 1]:
                    class_num.append(i)
                    break
                if int(time) < time_start[i] + 100 and int(time) + 100 > time_start[i + 1]:
                    class_num.append(i)
                    class_num.append(i + 1)
                    break
                if int(time) >= time_start[i] + 100 and int(time) + 100 > time_start[i + 1]:
                    class_num.append(i + 1)
                    break
    return class_num


def get_week(day):  # 获取活动设置在第几周
    for weekdays in week_date:
        if weekdays < 17:
            if int(day) > week_date[weekdays] and int(day) < week_date[weekdays + 1]:
                return weekdays
    else:
        return 0


def int_day(day):  # 将日期转为数字
    day_int = int(str(day)[0:4] + str(day)[5:7] + str(day)[8:10])
    return day_int


def IS_activitycrash(file_name, set_act_name, init_day, wday, IS_everyday, IS_everyweek, time_list):  # 查询活动与课程和其他活动是否冲突
    act_workbook = openpyxl.load_workbook(file_name)
    day_int = int_day(init_day)
    set_intweek = get_week(day_int)
    time_crash_num = {}
    if identity_data['class'] == 202101:
        wb = workbook2
    elif identity_data['class'] == 202102:
        wb = workbook3
    for start_time in time_list:
        time_crash_num[start_time] = 0
        for sheet in act_workbook:
            for i in range(2, sheet.max_row):
                if sheet.cell(row=i, column=1).value != set_act_name:
                    time = str(sheet.cell(row=i, column=6).value)[0:2] + str(sheet.cell(row=i, column=6).value)[3:5]
                    if int(start_time) >= int(time) - 100 and int(start_time) < int(time) + 100:
                        if sheet.cell(row=i, column=4).value == "Yes":
                            put_text(start_time, "与其他活动有冲突")
                            time_crash_num[start_time] += 1
                            put_text(sheet.cell(row=1, column=1).value, ":", sheet.cell(row=i, column=1).value)
                            put_text(sheet.cell(row=1, column=6).value, ":", sheet.cell(row=i, column=6).value)
                        elif sheet.cell(row=i, column=4).value == "No" and sheet.cell(row=i, column=5).value == "Yes":
                            if IS_everyday == "Yes":
                                put_text(start_time, "与其他活动有冲突")
                                time_crash_num[start_time] += 1
                                put_text(sheet.cell(row=1, column=1).value, ":", sheet.cell(row=i, column=1).value)
                                put_text(sheet.cell(row=1, column=6).value, ":", sheet.cell(row=i, column=6).value)
                            elif IS_everyday == "No" and IS_everyweek == "Yes":
                                if wday == sheet.cell(row=i, column=3).value:
                                    put_text(start_time, "与其他活动有冲突")
                                    time_crash_num[start_time] += 1
                                    put_text(sheet.cell(row=1, column=1).value, ":", sheet.cell(row=i, column=1).value)
                                    put_text(sheet.cell(row=1, column=6).value, ":", sheet.cell(row=i, column=6).value)
                            else:
                                if wday == sheet.cell(row=i, column=3).value:
                                    put_text(start_time, "与其他活动有冲突")
                                    time_crash_num[start_time] += 1
                                    put_text(sheet.cell(row=1, column=1).value, ":", sheet.cell(row=i, column=1).value)
                                    put_text(sheet.cell(row=1, column=6).value, ":", sheet.cell(row=i, column=6).value)
                        else:
                            com_day_int = int_day(sheet.cell(row=i, column=2).value)
                            if com_day_int > day_int:
                                if IS_everyday == "Yes":
                                    put_text(start_time, "与其他活动有冲突")
                                    time_crash_num[start_time] += 1
                                    put_text(sheet.cell(row=1, column=1).value, ":", sheet.cell(row=i, column=1).value)
                                    put_text(sheet.cell(row=1, column=6).value, ":", sheet.cell(row=i, column=6).value)
                                elif IS_everyday == "No" and IS_everyweek == "Yes":
                                    if wday == sheet.cell(row=i, column=3).value:
                                        put_text(start_time, "与其他活动有冲突")
                                        time_crash_num[start_time] += 1
                                        put_text(sheet.cell(row=1, column=1).value, ":",
                                                 sheet.cell(row=i, column=1).value)
                                        put_text(sheet.cell(row=1, column=6).value, ":",
                                                 sheet.cell(row=i, column=6).value)
        if int(start_time) >= 700 and int(start_time) <= 2000:
            classnum = get_crashclassnum(start_time)
            if len(classnum):
                if IS_everyday == "Yes":
                    for sheet_num in range(set_intweek, 16):
                        sheet = wb[wb.sheetnames[sheet_num]]
                        if sheet_num == set_intweek:
                            for ccolumn in range(wday + 1, 9):
                                for class_num in classnum:
                                    if sheet.cell(row=class_num + 1, column=ccolumn).value != "\\":
                                        put_text(start_time, "与课程",
                                                 sheet.cell(row=class_num + 1, column=ccolumn).value, "在第", sheet_num,
                                                 "周星期", ccolumn + 1, "第", class_num, "节课冲突")
                                        time_crash_num[start_time] += 1
                        else:
                            for ccolumn in range(2, 9):
                                for class_num in classnum:
                                    if sheet.cell(row=class_num + 1, column=ccolumn).value != "\\":
                                        put_text(start_time, "与课程",
                                                 sheet.cell(row=class_num + 1, column=ccolumn).value, "在第", sheet_num,
                                                 "周星期", ccolumn + 1, "第", class_num, "节课冲突")
                                        time_crash_num[start_time] += 1
                elif IS_everyday == "No" and IS_everyweek == "Yes":
                    for sheet_num in range(set_intweek - 1, 16):
                        sheet = wb[wb.sheetnames[sheet_num]]
                        for class_num in classnum:
                            if sheet.cell(row=class_num + 1, column=wday + 1).value != "\\":
                                put_text(start_time, "与课程", sheet.cell(row=class_num + 1, column=ccolumn).value,
                                         "在第", sheet_num, "周星期", ccolumn + 1, "第", class_num, "节课冲突")
                                time_crash_num[start_time] += 1
                elif IS_everyday == "No" and IS_everyweek == "No":
                    sheet = wb[wb.sheetnames[set_intweek - 1]]
                    for class_num in classnum:
                        if sheet.cell(row=class_num + 1, column=wday + 1).value != "\\":
                            put_text(start_time, "与课程", sheet.cell(row=class_num + 1, column=ccolumn).value, "在第",
                                     sheet_num, "周星期", ccolumn + 1, "第", class_num, "节课冲突")
                            time_crash_num[start_time] += 1
    return time_crash_num


def sort_activity(sort, file_name, rrow):  # 给活动排序
    activity = openpyxl.load_workbook(file_name)
    if sort == "A":
        sheet = activity[activity.sheetnames[0]]
    elif sort == "B":
        sheet = activity[activity.sheetnames[1]]
    set_time = str(sheet.cell(row=rrow, column=6).value)[0:2] + str(sheet.cell(row=rrow, column=6).value)[3:5]
    change_row = 0
    for row in range(2, sheet.max_row + 1):
        if row == rrow:
            continue
        elif row + 1 == rrow:
            next_row = row + 2
        else:
            next_row = row + 1
        if row < sheet.max_row:
            com_time = str(sheet.cell(row=row, column=6).value)[0:2] + str(sheet.cell(row=row, column=6).value)[3:5]
            if row == 2 and int(com_time) > int(set_time):
                change_row = row
                break
            if next_row <= sheet.max_row:
                com_time_next = str(sheet.cell(row=next_row, column=6).value)[0:2] + str(
                    sheet.cell(row=next_row, column=6).value)[3:5]
                if int(com_time) <= int(set_time) and int(com_time_next) > int(set_time):
                    change_row = row + 1
                    break
        if row == sheet.max_row:
            com_time = str(sheet.cell(row=row, column=6).value)[0:2] + str(sheet.cell(row=row, column=6).value)[3:5]
            if int(com_time) <= int(set_time):
                change_row = row + 1
                break
    if change_row != 0:
        sheet.insert_rows(change_row)
        if change_row < rrow:
            rrow += 1
        row_distance = change_row - rrow
        start_cell = sheet.cell(row=rrow, column=1).coordinate
        end_cell = sheet.cell(row=rrow, column=9).coordinate
        area = str(start_cell) + ":" + str(end_cell)
        sheet.move_range(area, rows=row_distance)
        sheet.delete_rows(rrow)
    activity.save(file_name)


def set_activity(current_real_time):  # 通过调用上面的函数实现对活动的设置
    i = 0
    for key in identity_data.keys():
        if key == 'stu_num':
            i += 1
    if i == 0:
        toast("不向非学生用户提供该服务")
        return None
    if identity_data['activity_clock_path'] == "\\":
        file_name = str(identity_data['name']) + ".xlsx"
        wb = openpyxl.Workbook()
        wb[wb.sheetnames[0]].title = "individual"
        wb.create_sheet(index=1, title="collective")
        wb['individual'].append(
            ['act_name', 'act_initday', 'act_initwday', 'IS_everyday', 'IS_everyweek', 'act_time', 'act_spot',
             'IS_online', 'website'])
        wb['collective'].append(
            ['act_name', 'act_initday', 'act_initwday', 'IS_everyday', 'IS_everyweek', 'act_time', 'act_spot',
             'IS_online', 'website'])
        wb.save(file_name)
        student_sheet.cell(row=int(student_row()), column=7).value = file_name
        workbook1.save(file_path1)
    else:
        file_name = identity_data['activity_clock_path']
    activity = openpyxl.load_workbook(file_name)
    sort = input("请输入类别，A.个人活动 B.集体活动")
    while sort != "A" and sort != "B":
        sort = input("请输入类别，A.个人活动 B.集体活动")
    if sort == "A":
        sheet = activity[activity.sheetnames[0]]
        data_df = pd.read_excel(file_name, sheet_name=None)
        data = data_df['individual']
    elif sort == "B":
        sheet = activity[activity.sheetnames[1]]
        data_df = pd.read_excel(file_name, sheet_name=None)
        data = data_df['collective']
    activity_name = input("请输入活动名称")
    for i in range(1, sheet.max_row + 2):
        if i <= sheet.max_row:
            if sheet.cell(row=i, column=1).value == activity_name:
                break
    while True:
        if i <= sheet.max_row:
            IS_new = 0
            break
        if i > sheet.max_row:
            IS_new = input("是否为新增活动,输入1是，Exit退出设置，其他不是")
            if IS_new == "1":
                break
            elif IS_new == "Exit":
                return None
        activity_name = input("请输入活动名称")
        for i in range(1, sheet.max_row + 2):
            if i <= sheet.max_row:
                if sheet.cell(row=i, column=1).value == activity_name:
                    break
    if IS_new == 0:
        toast("有该活动")
        with popup(title="详情"):
            put_text(data.loc[i - 2])
        logging.info('用户更改%s活动', str(activity_name))
        j = 2
        IS_false = 0
        act_info = []
        for k in range(2, sheet.max_column + 1):
            act_info.append(sheet.cell(row=i, column=k).value)
        while j <= sheet.max_column:
            if j != 3:
                toast(sheet.cell(row=1, column=j).value)
            if j == 2:
                IS_in = input("是否设置初始日期，回车进入下一项，其他任意键进如修改")
                if IS_in == "":
                    sheet.cell(row=i, column=j).value = current_real_time.strftime("%Y-%m-%d")
                    toast(sheet.cell(row=i, column=2).value)
                    logging.info('用户更改%s活动的初始日期为%s', str(activity_name),
                                 str(sheet.cell(row=i, column=j).value))
                    activity.save(file_name)
            elif j > 2 and j != 3 and j != 5 and j != 9:
                IS_in = input("是否进入修改本项，回车进入下一项，-1进入上一级修改，其他任意键进如修改")
            elif j == 5:
                if sheet.cell(row=i, column=4).value == "No":
                    IS_in = input("活动不是每天都要进行，回车表示不需要重新设置周频率,其他键表示需要,-1回上一级")
                elif sheet.cell(row=i, column=4).value == "Yes":
                    IS_reback = input("活动每天都进行，已自动将周频率关掉,-1表示回到上一级,其他进入下一级设置")
                    if IS_reback == "-1":
                        j -= 1
                        continue
                    else:
                        sheet.cell(row=i, column=j).value = "No"
                        activity.save(file_name)
                        logging.info('用户更改%s活动，不是每周进行', str(activity_name))
                        j += 1
                        continue
            elif j == 9:
                if sheet.cell(row=i, column=8).value == "Yes":
                    IS_in = input("线上活动，回车表示不需要重新设置活动链接，其他键表示需要")
                    if IS_in != "":
                        IS_in = '0'
                elif sheet.cell(row=i, column=8).value == "No":
                    toast("线下活动，已自动将网站设为空")
                    sheet.cell(row=i, column=j).value = "\\"
                    logging.info('用户更改的%s活动为线下活动，空链接', str(activity_name))
                    activity.save(file_name)
                    j += 1
                    continue
            if IS_in == "":
                j += 1
                continue
            elif IS_in == '-1' and j > 2 and j != 4:
                j -= 1
                continue
            elif IS_in == '-1' and j == 4:
                j -= 2
                continue
            if sheet.cell(row=1, column=j).value == "act_initday":
                toast("请依次按照格式输入年月日")
                year = input("年----")
                month = input("月--")
                day = input("日--")
                sheet.cell(row=i, column=j).value = str(year) + "-" + str(month) + "-" + str(day)
                logging.info('用户更改%s活动的初始日期为%s', str(activity_name), str(sheet.cell(row=i, column=j).value))
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "act_initwday":
                date_str = str(sheet.cell(row=i, column=2).value)
                sheet.cell(row=i, column=j).value = time.strptime(date_str, '%Y-%m-%d').tm_wday + 1
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "IS_everyday":
                info = input("是否每天进行该活动Yes or No")
                while info != "Yes" and info != "No":
                    info = input("是否每天进行该活动Yes or No")
                sheet.cell(row=i, column=j).value = info
                activity.save(file_name)
                if info == 'Yes':
                    logging.info('用户更改的%s活动为每天进行', str(activity_name))
                elif info == 'No':
                    logging.info('用户更改的%s活动为不是每天进行', str(activity_name))
            if sheet.cell(row=1, column=j).value == "IS_everyweek":
                info = input("是否每周进行Yes or No")
                while info != "Yes" and info != "No":
                    info = input("是否每周进行Yes or No")
                sheet.cell(row=i, column=j).value = info
                activity.save(file_name)
                if info == 'Yes':
                    logging.info('用户更改的%s活动为每周进行', str(activity_name))
                elif info == 'No':
                    logging.info('用户更改的%s活动为不是每周进行', str(activity_name))
            if sheet.cell(row=1, column=j).value == "act_time":
                time_list = []
                while len(time_list) == 0:
                    info = '0'
                    while info != '#':
                        info = input(
                            "输入活动开始时间是几点，有效时间从6点到22点--：--,如果输入#结束，但至少需要输入一个时间")
                        while (info.isdigit() != True and info != '#') or (
                                info.isdigit() and (int(info) < 600 or int(info) > 2200)):
                            info = input("输入无效，请重新输入,如果输入#结束，但至少需要输入一个时间")
                        if info != '#':
                            time_list.append(info)
                init_day = sheet.cell(row=i, column=2).value
                IS_everyday = sheet.cell(row=i, column=4).value
                IS_everyweek = sheet.cell(row=i, column=5).value
                init_wday = sheet.cell(row=i, column=3).value
                time_crash_num = IS_activitycrash(file_name, activity_name, init_day, init_wday, IS_everyday,
                                                  IS_everyweek, time_list)
                if sort == 'A':
                    R_time = 0
                    for time_info in time_crash_num:
                        if time_crash_num[time_info] == 0:
                            R_time = time_info
                    if R_time == 0:
                        toast("输出失败")
                        break
                if sort == 'B':
                    R_time = 0
                    crash_num = sorted(time_crash_num.items(), key=lambda x: x[0])
                    with popup(title='冲突最少的事件(冲突的次数)'):
                        put_text(crash_num[0])
                    R_time = crash_num[0][0]
                sheet.cell(row=i, column=j).value = str(R_time)[0:2] + ":" + str(R_time)[2:4]
                logging.info('用户更改的%s活动时间%s', str(activity_name), str(sheet.cell(row=i, column=j).value))
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "act_spot":
                with popup("所有地点"):
                    for fra in infrastructure_data:
                        put_text(fra['infra_name'])
                info = input("请输入活动地点")
                for infra in range(1, infra_sheet.max_row + 2):
                    if infra <= infra_sheet.max_row:
                        if infra_sheet.cell(row=infra, column=2).value == info:
                            sheet.cell(row=i, column=j).value = info
                            break
                while infra > infra_sheet.max_row:
                    info = input("请重新输入：")
                    for infra in range(1, infra_sheet.max_row + 2):
                        if infra <= infra_sheet.max_row:
                            if infra_sheet.cell(row=infra, column=2).value == info:
                                sheet.cell(row=i, column=j).value = info
                                break
                logging.info('用户更改的%s活动地点为%s', str(activity_name), str(sheet.cell(row=i, column=j).value))
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "IS_online":
                info = input("是否线上Yes or No")
                while info != "Yes" and info != "No":
                    info = input("是否线上Yes or No")
                sheet.cell(row=i, column=j).value = info
                if info == 'Yes':
                    logging.info('用户更改的%s活动为线上进行', str(activity_name))
                    sheet.cell(row=i, column=j - 1).value = '\\'
                    logging.info('用户更改的%s活动为线上进行,上课地点改为空', str(activity_name))
                elif info == 'No':
                    logging.info('用户更改的%s活动为线下进行', str(activity_name))
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "website":
                info = input("请输入活动链接")
                sheet.cell(row=i, column=j).value = info
                logging.info('用户更改的%s活动为线上动，链接为%s', str(activity_name), str(info))
                activity.save(file_name)
                IS_reback = input("1表示重新设置链接，-1回到上一级设置，回车结束")
                if IS_reback == '1':
                    IS_in = '0'
                    break
                if IS_reback == '-1':
                    IS_in = '0'
                    j -= 1
                    break
            j += 1
        if IS_false:
            logging.info('用户更改的%s活动失败', str(activity_name))
            for k in range(2, sheet.max_column + 1):
                sheet.cell(row=i, column=k).value = act_info[k - 2]
            activity.save(file_name)
    elif IS_new == "1":
        toast("新增活动")
        j = 1
        IS_false = 0
        while j <= sheet.max_column:
            if j > 2:
                if IS_in == '-1' and j > 2 and j != 5 and j != 9:
                    j -= 2
                elif IS_in == "-1" and j == 5:
                    j -= 3
                elif IS_in == "1":
                    j -= 1
            if j > 1 and j != 3:
                toast(sheet.cell(row=1, column=j).value)
            if j == 1:
                sheet.cell(row=i, column=j).value = activity_name
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "act_initday":
                IS_today = input("输入1表示默认为今天,其他表示初始日期自己设置")
                if IS_today == '1':
                    sheet.cell(row=i, column=j).value = current_real_time.strftime("%Y-%m-%d")
                    toast(sheet.cell(row=i, column=2).value)
                else:
                    toast("请依次按照格式输入年月日")
                    year = input("年----")
                    month = input("月--")
                    day = input("日--")
                    sheet.cell(row=i, column=j).value = str(year) + "-" + str(month) + "-" + str(day)
                logging.info('用户设置%s活动的初始日期为%s', str(activity_name), str(sheet.cell(row=i, column=j).value))
                activity.save(file_name)
                IS_in = input("输入1表示需要重新修改本项，其他任意键进入下一项设置")
            if sheet.cell(row=1, column=j).value == "act_initwday":
                date_str = str(sheet.cell(row=i, column=2).value)
                sheet.cell(row=i, column=j).value = time.strptime(date_str, '%Y-%m-%d').tm_wday + 1
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "IS_everyday":
                info = input("活动是否每天进行Yes or No")
                while info != "Yes" and info != "No":
                    info = input("活动是否每天进行Yes or No")
                sheet.cell(row=i, column=j).value = info
                activity.save(file_name)
                if info == 'Yes':
                    logging.info('用户设置的%s活动为每天进行', str(activity_name))
                elif info == 'No':
                    logging.info('用户设置的%s活动为不是每天进行', str(activity_name))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            if sheet.cell(row=1, column=j).value == "IS_everyweek":
                if sheet.cell(row=i, column=4).value == "No":
                    info = input("活动不是每天进行，是否每周进行Yes or No")
                    while info != "Yes" and info != "No":
                        info = input("活动不是每天进行，是否每周进行Yes or No")
                    sheet.cell(row=i, column=j).value = info
                else:
                    toast("活动每天都进行，自动将周频率关掉")
                    sheet.cell(row=i, column=j).value = "No"
                activity.save(file_name)
                if info == 'Yes':
                    logging.info('用户设置的%s活动为每周进行', str(activity_name))
                elif info == 'No':
                    logging.info('用户设置的%s活动为不是每周进行', str(activity_name))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            if sheet.cell(row=1, column=j).value == "act_time":
                time_list = []
                while len(time_list) == 0:
                    info = '0'
                    while info != '#':
                        info = input(
                            "输入活动开始时间是几点，有效时间从6点到22点--：--,如果输入#结束，但至少需要输入一个时间")
                        while (info.isdigit() != True and info != '#') or (
                                info.isdigit() and (int(info) < 600 or int(info) > 2200)):
                            info = input("输入无效，请重新输入,如果输入#结束，但至少需要输入一个时间")
                        if info != '#':
                            time_list.append(info)
                init_day = sheet.cell(row=i, column=2).value
                IS_everyday = sheet.cell(row=i, column=4).value
                IS_everyweek = sheet.cell(row=i, column=5).value
                init_wday = sheet.cell(row=i, column=3).value
                time_crash_num = IS_activitycrash(file_name, activity_name, init_day, init_wday, IS_everyday,
                                                  IS_everyweek, time_list)
                if sort == 'A':
                    R_time = 0
                    for time_info in time_crash_num:
                        if time_crash_num[time_info] == 0:
                            R_time = time_info
                    if R_time == 0:
                        IS_in = input("时间设置都有冲突，1表示重新设置时间,-1表示退回上一级设置，回车设置失败")
                        if IS_in == "":
                            IS_false = 1
                            break
                if sort == 'B':
                    R_time = 0
                    crash_num = sorted(time_crash_num.items(), key=lambda x: x[0])
                    put_text('冲突最少的事件(冲突的次数)', crash_num[0])
                    R_time = crash_num[0][0]
                    IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
                sheet.cell(row=i, column=j).value = str(R_time)[0:2] + ":" + str(R_time)[2:4]
                logging.info('用户更改的%s活动时间%s', str(activity_name), str(sheet.cell(row=i, column=j).value))
                activity.save(file_name)
            if sheet.cell(row=1, column=j).value == "act_spot":
                with popup("所有地点"):
                    for fra in infrastructure_data:
                        put_text(fra['infra_name'])
                info = input("请输入活动地点")
                for infra in range(1, infra_sheet.max_row + 2):
                    if infra <= infra_sheet.max_row:
                        if infra_sheet.cell(row=infra, column=2).value == info:
                            sheet.cell(row=i, column=j).value = info
                            break
                while infra > infra_sheet.max_row:
                    info = input("请重新输入：")
                    for infra in range(1, infra_sheet.max_row + 2):
                        if infra <= infra_sheet.max_row:
                            if infra_sheet.cell(row=infra, column=2).value == info:
                                sheet.cell(row=i, column=j).value = info
                                break
                activity.save(file_name)
                logging.info('用户设置的%s活动地点为%s', str(activity_name), str(sheet.cell(row=i, column=j).value))
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            if sheet.cell(row=1, column=j).value == "IS_online":
                info = input("活动是否线上Yes or No")
                while info != "Yes" and info != "No":
                    info = input("活动是否线上Yes or No")
                sheet.cell(row=i, column=j).value = info
                if info == 'Yes':
                    logging.info('用户设置的%s活动为线上进行', str(activity_name))
                    sheet.cell(row=i, column=j - 1).value = '\\'
                    logging.info('用户设置的%s活动为线上进行,上课地点改为空', str(activity_name))
                elif info == 'No':
                    logging.info('用户设置的%s活动为线下进行', str(activity_name))
                activity.save(file_name)
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键进入下一项设置")
            if sheet.cell(row=1, column=j).value == "website":
                if sheet.cell(row=i, column=8).value == "Yes":
                    info = input("活动是线上活动，请输入活动链接")
                    sheet.cell(row=i, column=j).value = info
                    logging.info('用户设置的%s活动为线上活动，链接为空', str(activity_name))
                else:
                    toast("活动是线下，已自动将活动链接设为空")
                    sheet.cell(row=i, column=j).value = "\\"
                    logging.info('用户设置的%s活动为线下活动，链接为%s', str(activity_name), str(info))
                activity.save(file_name)
                IS_in = input("输入1表示需要重新修改本项，-1表示修改上一项，其他任意键表示结束")
                if IS_in == "1":
                    j -= 1
                elif IS_in == '-1':
                    j -= 2
                IS_in = '0'
            j += 1
        if IS_false:
            sheet.delete_rows(i)
            logging.info('用户设置的%s活动失败', str(activity_name))
            activity.save(file_name)
    if IS_false != 1:
        sort_activity(sort, file_name, i)
    return file_name


def query_act():  # 查询活动
    i = 0
    for key in identity_data.keys():
        if key == 'stu_num':
            i += 1
    if i == 0:
        toast("不向非学生用户提供该服务")
        return None
    if identity_data['activity_clock_path'] == "\\":
        toast("你还没有活动")
        return None
    act = pd.read_excel(identity_data['activity_clock_path'], sheet_name=None)
    indi_data = act['individual'].to_dict(orient='records')
    col_data = act['collective'].to_dict(orient='records')
    count = 0
    act_info = input("请输入你要查询的活动信息，时间形如--:--例如0823表示08：23")
    if act_info.isdigit():
        for indi in indi_data:
            if int(act_info) >= int(str(indi['act_time'])[0:2] + str(indi['act_time'])[3:5]) and int(act_info) <= int(
                    str(indi['act_time'])[0:2] + str(indi['act_time'])[3:5]) + 100:
                with use_scope('scope2', clear=True):
                    put_markdown('#### 个人活动')
                    put_text(indi)
                count += 1
        for col in col_data:
            if int(act_info) >= int(str(col['act_time'])[0:2] + str(col['act_time'])[3:5]) and int(act_info) <= int(
                    str(col['act_time'])[0:2] + str(col['act_time'])[3:5]) + 100:
                with use_scope('scope1', clear=True):
                    put_markdown('#### 集体活动')
                    put_text(col)
                count += 1
        logging.info('用户查询%s时的活动', str(act_info)[0:2] + ":" + str(act_info)[2:4])
    else:
        for indi in indi_data:
            if indi['act_name'] == act_info:
                with use_scope('scope2', clear=True):
                    put_markdown('#### 个人活动')
                    put_text(indi)
                count += 1
        for col in col_data:
            if col['act_name'] == act_info:
                with use_scope('scope1', clear=True):
                    put_markdown('#### 集体活动')
                    put_text(col)
                count += 1
        logging.info('用户更改%s活动', str(act_info))
    if count == 0:
        toast("查询失败")
    sort = input("是否查询类别，A查询所有个人活动 B.查询所有集体活动 C.查询所有活动 D.不需要查询")
    if sort == "A":
        if act["individual"].shape[0] == 0:
            toast("没有个人活动")
        else:
            with use_scope('scope2', clear=True):
                put_markdown('#### 个人活动')
                put_text(act['individual'])
        logging.info('用户查询所有个人活动')
    elif sort == "B":
        if act['collective'].shape[0] == 0:
            toast("没有集体活动")
        else:
            with use_scope('scope1', clear=True):
                put_markdown('#### 集体活动')
                put_text(act['collective'])
        logging.info('用户查询所有集体活动')
    elif sort == "C":
        if act["individual"].shape[0] == 0:
            toast("没有个人活动")
        else:
            with use_scope('scope2', clear=True):
                put_markdown('#### 个人活动')
                put_text(act['individual'])
        if act['collective'].shape[0] == 0:
            toast("没有集体活动")
        else:
            with use_scope('scope1', clear=True):
                put_markdown('#### 集体活动')
                put_text(act['collective'])
        logging.info('用户查询所有活动')


def clock_ring(current_real_time, identity_data):  # 每天晚上输出第二天的课程信息
    second_All = ''

    act_filename = identity_data['activity_clock_path']
    global second_day_data
    time_data = str(current_real_time.strftime('%H')) + '00'  # 当天时间
    date = str(current_real_time.strftime('%Y%m%d'))
    if int(time_data) == 2300:
        second_day_data = {}
        week = int(get_week(date))
        if time.strptime(date, '%Y%m%d').tm_wday + 1 == 7:
            wday = 1
            week += 1
        else:
            wday = time.strptime(date, '%Y%m%d').tm_wday + 2
        if identity_data['class'] == 202101:
            wb = workbook2
        elif identity_data['class'] == 202102:
            wb = workbook3
        sheet = wb[wb.sheetnames[week - 1]]
        for i in range(2, sheet.max_row + 1):
            course = sheet.cell(row=i, column=wday + 1).value
            IS_exam = 0
            if course != "\\":
                if str(course)[-5:] == ".exam":
                    course = str(course)[:-5]
                    IS_exam = 1
                for course_info in course_data:
                    if course_info['course_name'] == course and course_info['class'] == identity_data['class']:
                        if IS_exam == 0:
                            data_name = str(course_info['course_name'])
                            second_day_data[data_name] = {}
                            second_day_data[data_name]['名称'] = data_name
                            second_day_data[data_name]['time'] = time_start_str[i - 1]
                            second_day_data[data_name]['spot'] = course_info['course_spot']
                            second_day_data[data_name]['online'] = course_info['online']
                            second_day_data[data_name]['website'] = course_info['website']
                        elif IS_exam == 1:
                            data_name = str(course_info['course_name']) + ".exam"
                            second_day_data[data_name] = {}
                            second_day_data[data_name]['名称'] = data_name
                            second_day_data[data_name]['time'] = time_start_str[i - 1]
                            second_day_data[data_name]['spot'] = course_info['exam_spot']
                            second_day_data[data_name]['online'] = '\\'
                            second_day_data[data_name]['website'] = '\\'
        if act_filename != '\\':
            activity = openpyxl.load_workbook(act_filename)
            for act_sheet in activity:
                for i in range(2, act_sheet.max_row + 1):
                    if act_sheet.cell(row=i, column=4).value == 'Yes':
                        data_name = str(act_sheet.cell(row=i, column=1).value)
                        second_day_data[data_name] = {}
                        second_day_data[data_name]['名称'] = data_name
                        second_day_data[data_name]['time'] = act_sheet.cell(row=i, column=6).value
                        second_day_data[data_name]['spot'] = act_sheet.cell(row=i, column=7).value
                        second_day_data[data_name]['online'] = act_sheet.cell(row=i, column=8).value
                        second_day_data[data_name]['website'] = act_sheet.cell(row=i, column=9).value
                    elif act_sheet.cell(row=i, column=4).value == 'No' and act_sheet.cell(row=i,
                                                                                          column=5).value == 'Yes':
                        if int(act_sheet.cell(row=i, column=3).value) == wday:
                            data_name = str(act_sheet.cell(row=i, column=1).value)
                            second_day_data[data_name] = {}
                            second_day_data[data_name]['名称'] = data_name
                            second_day_data[data_name]['time'] = act_sheet.cell(row=i, column=6).value
                            second_day_data[data_name]['spot'] = act_sheet.cell(row=i, column=7).value
                            second_day_data[data_name]['online'] = act_sheet.cell(row=i, column=8).value
                            second_day_data[data_name]['website'] = act_sheet.cell(row=i, column=9).value
                    else:
                        com_date = int(int_day(act_sheet.cell(row=i, column=2).value))
                        if com_date == int(date) + 1:
                            data_name = str(act_sheet.cell(row=i, column=1).value)
                            second_day_data[data_name] = {}
                            second_day_data[data_name]['名称'] = data_name
                            second_day_data[data_name]['time'] = act_sheet.cell(row=i, column=6).value
                            second_day_data[data_name]['spot'] = act_sheet.cell(row=i, column=7).value
                            second_day_data[data_name]['online'] = act_sheet.cell(row=i, column=8).value
                            second_day_data[data_name]['website'] = act_sheet.cell(row=i, column=9).value
        for data in second_day_data:
            second_All += str(second_day_data[data]) + '\n'

        toast('明天的课程 🔔')
        popup('空则代表没有', second_All)


def get_destination(current_real_time):
    destination = []
    time_data = str(current_real_time.strftime('%H')) + '00'
    x = 0
    y = 0
    for data in second_day_data:
        if int(str(second_day_data[data]['time'])[0:2] + str(second_day_data[data]['time'])[3:5]) >= int(
                time_data) and int(
                str(second_day_data[data]['time'])[0:2] + str(second_day_data[data]['time'])[3:5]) < int(
                time_data) + 100:
            with use_scope('class_next', clear=True):
                put_markdown('#### 最近的活动或课程')
                put_text(data, second_day_data[data])
                if second_day_data[data]['spot'] != '\\':
                    infra = second_day_data[data]['spot']
                    for infrastructure in infrastructure_data:
                        if infrastructure['infra_name'] == infra:
                            x = infrastructure['x']
                            y = infrastructure['y']
                            destination.append([x, y])
                            break
    if len(destination) == 0:
        destination.append([0, 0])
    return destination  # 如果[x,y]==[0,0]说明是线上的课程或活动


def get_todaydata(current_real_time, identity_data):
    act_filename = identity_data['activity_clock_path']
    date = str(current_real_time.strftime('%Y%m%d'))
    wday = time.strptime(date, '%Y%m%d').tm_wday + 1
    week = int(get_week(date))
    global second_day_data
    second_day_data = {}
    if identity_data['class'] == 202101:
        wb = workbook2
    elif identity_data['class'] == 202102:
        wb = workbook3
    sheet = wb[wb.sheetnames[week - 1]]
    for i in range(2, sheet.max_row + 1):
        course = sheet.cell(row=i, column=wday + 1).value
        IS_exam = 0
        if course != "\\":
            if str(course)[-5:] == ".exam":
                course = str(course)[:-5]
                IS_exam = 1
            for course_info in course_data:
                if course_info['course_name'] == course and course_info['class'] == identity_data['class']:
                    if IS_exam == 0:
                        data_name = str(course_info['course_name'])
                        second_day_data[data_name] = {}
                        second_day_data[data_name]['time'] = time_start_str[i - 1]
                        second_day_data[data_name]['spot'] = course_info['course_spot']
                        second_day_data[data_name]['online'] = course_info['online']
                        second_day_data[data_name]['website'] = course_info['website']
                    elif IS_exam == 1:
                        data_name = str(course_info['course_name']) + ".exam"
                        second_day_data[data_name] = {}
                        second_day_data[data_name]['time'] = time_start_str[i - 1]
                        second_day_data[data_name]['spot'] = course_info['exam_spot']
                        second_day_data[data_name]['online'] = '\\'
                        second_day_data[data_name]['website'] = '\\'
    if act_filename != '\\':
        activity = openpyxl.load_workbook(act_filename)
        for act_sheet in activity:
            for i in range(2, act_sheet.max_row + 1):
                if act_sheet.cell(row=i, column=4).value == 'Yes':
                    data_name = str(act_sheet.cell(row=i, column=1).value)
                    second_day_data[data_name] = {}
                    second_day_data[data_name]['time'] = act_sheet.cell(row=i, column=6).value
                    second_day_data[data_name]['spot'] = act_sheet.cell(row=i, column=7).value
                    second_day_data[data_name]['online'] = act_sheet.cell(row=i, column=8).value
                    second_day_data[data_name]['website'] = act_sheet.cell(row=i, column=9).value
                elif act_sheet.cell(row=i, column=4).value == 'No' and act_sheet.cell(row=i, column=5).value == 'Yes':
                    if int(act_sheet.cell(row=i, column=3).value) == wday:
                        data_name = str(act_sheet.cell(row=i, column=1).value)
                        second_day_data[data_name] = {}
                        second_day_data[data_name]['time'] = str(act_sheet.cell(row=i, column=6).value)[0:2] + str(
                            act_sheet.cell(row=i, column=6).value)[3:5]
                        second_day_data[data_name]['spot'] = act_sheet.cell(row=i, column=7).value
                        second_day_data[data_name]['online'] = act_sheet.cell(row=i, column=8).value
                        second_day_data[data_name]['website'] = act_sheet.cell(row=i, column=9).value
                else:
                    com_date = int(int_day(act_sheet.cell(row=i, column=2).value))
                    if com_date == int(date) + 1:
                        data_name = str(act_sheet.cell(row=i, column=1).value)
                        second_day_data[data_name] = {}
                        second_day_data[data_name]['time'] = str(act_sheet.cell(row=i, column=6).value)[0:2] + str(
                            act_sheet.cell(row=i, column=6).value)[3:5]
                        second_day_data[data_name]['spot'] = act_sheet.cell(row=i, column=7).value
                        second_day_data[data_name]['online'] = act_sheet.cell(row=i, column=8).value
                        second_day_data[data_name]['website'] = act_sheet.cell(row=i, column=9).value


# 主操作模块


def course_init(current_real_time, identity_this):
    import sys
    sys.stdin.flush()
    global identity_data
    identity_data = identity_this  # 如果是学生，获得该学生的班级，如果是辅导员，获取辅导员的账号，不要再改了！已经按照变量名不同的逻辑写了
    # 因为没有任何能区分的通用变量名，所有分两种情况
    global act_filename
    if identity_data['if_stu'] == 1:
        # 个人日志文件路径
        act_filename = identity_data['activity_clock_path']
        stu_num = identity_data['stu_num']
        log_file_foruse = log_file + "Log-" + str(stu_num) + ".txt"
        logging.basicConfig(filename=log_file_foruse, level=logging.DEBUG, format='%(asctime)s -- %(message)s')
    elif identity_data['if_stu'] == 0:
        ins_account = identity_data['ins_account']
        log_file_foruse = log_file + "Log-" + str(ins_account) + ".txt"
        logging.basicConfig(filename=log_file_foruse, level=logging.DEBUG, format='%(asctime)s -- %(message)s')

    if identity_data != 0:
        infunc = input("查询课表1，更改课程表2，查询活动3，更改活动和活动闹钟4，其他输入退出，输入想完成的操作")
        if infunc.isdigit():
            if int(infunc) == 1:
                toast("查询课表")
                querycourse()

            elif int(infunc) == 2:
                toast("更改课程")
                set_timetable()

            elif int(infunc) == 3:
                toast("查询活动")
                query_act()

            elif int(infunc) == 4:
                toast("更改活动")
                act_filename = set_activity(current_real_time)

# course_init()